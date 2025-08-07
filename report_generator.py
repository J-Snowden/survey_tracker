import pandas as pd
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ReportGenerator:
    def __init__(self, report_dir="reports"):
        self.report_dir = report_dir
        self.logger = logging.getLogger(__name__)
        
        # Ensure report directory exists
        os.makedirs(self.report_dir, exist_ok=True)
        
    def generate_report(self, data, filename=None, teacher_data=None):
        """
        Generate an Excel report from the processed data
        
        Args:
            data (dict): Aggregated data with teacher IDs, dates, and response counts
            filename (str): Optional filename for the report
            teacher_data (list): Optional teacher tracking data for second tab
            
        Returns:
            str: Path to the generated report file
        """
        try:
            # Create DataFrame from data
            if not data:
                self.logger.warning("No data to generate report from")
                # Create empty DataFrame with proper columns
                df = pd.DataFrame(columns=["Teacher ID", "Date", "Number of Responses", "Number of Blanks"])
            else:
                # Convert data to DataFrame
                report_data = []
                for (teacher_id, date), counts in data.items():
                    report_data.append({
                        "Teacher ID": teacher_id,
                        "Date": date,
                        "Time": counts.get("time_type", ""),
                        "Filename": counts.get("filename", ""),
                        "Number of Responses": counts["responses"],
                        "Number of Blanks": counts["blanks"]
                    })
                
                df = pd.DataFrame(report_data)
                
                # Sort by Teacher ID and Date
                df = df.sort_values(["Teacher ID", "Date"])
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"survey_report_{timestamp}.xlsx"
                
            # Create full file path
            file_path = os.path.join(self.report_dir, filename)
            
            # Generate Excel report
            self._create_excel_report(df, file_path, teacher_data)
            
            self.logger.info(f"Report generated successfully: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}")
            raise
    
    def _create_excel_report(self, df, file_path, teacher_data=None):
        """
        Create an Excel report with formatting
        
        Args:
            df (pandas.DataFrame): DataFrame with report data
            file_path (str): Path to save the Excel file
            teacher_data (list): Optional teacher tracking data for second tab
        """
        try:
            # Create a new workbook and select the active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Survey Report"
            
            # Add title row
            title_cell = ws.cell(row=1, column=1, value="Survey Response Report")
            title_cell.font = Font(size=16, bold=True)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            
            # Add generation date
            date_cell = ws.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            date_cell.font = Font(italic=True)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
            ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
            
            # Add headers (row 4)
            headers = ["Teacher ID", "Date", "Time", "Filename", "Number of Responses", "Number of Blanks"]
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 5):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                # Get the column letter from the first cell that is not merged
                column_letter = None
                for cell in column:
                    try:
                        # Check if the cell has a column_letter attribute (not merged)
                        if hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                            break
                    except:
                        pass
                
                # If we couldn't find a cell with column_letter, skip this column
                if column_letter is None:
                    continue
                    
                # Calculate max length for this column
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
            # Create teacher summary tab if teacher data is provided
            if teacher_data:
                self._create_teacher_summary_tab(wb, teacher_data)
            
            # Save the workbook
            wb.save(file_path)
            
        except Exception as e:
            self.logger.error(f"Error creating Excel report: {str(e)}")
            raise
    
    def generate_summary_statistics(self, data):
        """
        Generate summary statistics from the data
        
        Args:
            data (dict): Aggregated data with teacher IDs, dates, and response counts
            
        Returns:
            dict: Summary statistics
        """
        if not data:
            return {
                "total_teachers": 0,
                "total_responses": 0,
                "date_range": "No data",
                "avg_responses_per_teacher": 0
            }
        
        # Extract unique teachers and dates
        teachers = set()
        dates = set()
        total_responses = 0
        total_blanks = 0
        
        for (teacher_id, date), counts in data.items():
            teachers.add(teacher_id)
            dates.add(date)
            total_responses += counts["responses"]
            total_blanks += counts["blanks"]
            
        # Calculate date range
        if dates:
            sorted_dates = sorted(dates)
            date_range = f"{sorted_dates[0]} to {sorted_dates[-1]}"
        else:
            date_range = "No dates"
            
        # Calculate average responses per teacher
        avg_responses = total_responses / len(teachers) if teachers else 0
        
        return {
            "total_teachers": len(teachers),
            "total_responses": total_responses,
            "total_blanks": total_blanks,
            "date_range": date_range,
            "avg_responses_per_teacher": round(avg_responses, 2)
        }
    
    def _create_teacher_summary_tab(self, workbook, teacher_data):
        """
        Create the Teacher Summary tab in the workbook
        
        Args:
            workbook: The openpyxl workbook object
            teacher_data (list): List of teacher statistics dictionaries
        """
        try:
            # Create new worksheet
            ws = workbook.create_sheet(title="Teacher Summary")
            
            # Add title row
            title_cell = ws.cell(row=1, column=1, value="Teacher Summary Report")
            title_cell.font = Font(size=16, bold=True)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            
            # Add generation date
            date_cell = ws.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            date_cell.font = Font(italic=True)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
            ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
            
            # Add headers (row 4)
            headers = ["Teacher ID", "Pre Responses", "First Pre Date", "Last Pre Date",
                      "Post Responses", "First Post Date", "Last Post Date"]
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_idx, teacher_record in enumerate(teacher_data, 5):
                ws.cell(row=row_idx, column=1, value=teacher_record['Teacher ID'])
                # Show blank instead of 0 for responses
                pre_responses = teacher_record['Pre Responses'] if teacher_record['Pre Responses'] > 0 else ""
                post_responses = teacher_record['Post Responses'] if teacher_record['Post Responses'] > 0 else ""
                ws.cell(row=row_idx, column=2, value=pre_responses)
                ws.cell(row=row_idx, column=3, value=teacher_record['First Pre Date'])
                ws.cell(row=row_idx, column=4, value=teacher_record['Last Pre Date'])
                ws.cell(row=row_idx, column=5, value=post_responses)
                ws.cell(row=row_idx, column=6, value=teacher_record['First Post Date'])
                ws.cell(row=row_idx, column=7, value=teacher_record['Last Post Date'])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = None
                
                for cell in column:
                    try:
                        if hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                            break
                    except:
                        pass
                
                if column_letter is None:
                    continue
                    
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = min(adjusted_width, 25)
                
        except Exception as e:
            self.logger.error(f"Error creating teacher summary tab: {str(e)}")
            raise